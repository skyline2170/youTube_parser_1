import cProfile
import csv
import datetime
import multiprocessing
import os.path
import time

import openpyxl

import urllib3

import fake_useragent
import requests as req

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from tqdm import tqdm
import json_validator


class YouTube_Parser:

    def __init__(self, url: str, multiproc=False):
        self.run_check = 0
        self.__multiproc = multiproc
        self.__chanal_url = url
        self.__user_agent = fake_useragent.UserAgent()
        self.__driver = None
        self.href_list = []
        self.all_video_data_list = []
        self.lost_href_list = []
        self.page_source = None

    def page_scroller(self):
        '''Даная функция пролистывает интернет страницу, открытую в драйвере, до самого низа. Основано на подсчёте
        количества ссылок на странице. Если кол-во ссылок не меняется с пролистыванием, то страница пролистана.'''
        start = datetime.datetime.now()
        k = 0  # старое количество страниц
        check = 0
        while True:  # цикл пролистывает страницу до конца
            # time.sleep(3)
            time.sleep(0.5)
            hrefs = self.__driver.find_elements(By.TAG_NAME, "a")
            if len(hrefs) == k:
                check += 1
                if check > 3:
                    print("Поиск видео окончен.")
                    break
            else:
                now = datetime.datetime.now()
                if str(now - start) > "0:00:15.000000":
                    start = now
                    print("Поиск продоложается...")
                k = len(hrefs)
                self.__driver.execute_script("window.scrollBy(0,1000000);")

    def get_hrefs(self):
        ''''Данныя функция получает все ссылки на видео со страницы, открытой в драйвере, и сохраняет их в self.href_list'''
        print("Получение ссылок...")
        html_page = self.__driver.find_element(By.XPATH, '//*[@id="contents"]')  # находим div блок с видео
        # html_page = html_page.find_element(By.TAG_NAME, 'div')
        # print(f"{html_page.get_attribute('id')}")
        html_page = html_page.find_elements(By.ID, "details")

        # html_page = html_page.find_elements(By.TAG_NAME, "h3")
        # print(html_page)
        self.href_list = [i.find_element(By.TAG_NAME, "h3") for i in html_page]
        self.href_list = [i.find_element(By.TAG_NAME, "a") for i in html_page]
        # print("href[0]", self.href_list[0].get_attribute("href"))

        if self.href_list[0].get_attribute("href") == None:
            return None
        # print(self.href_list)
        # self.href_list = html_page.find_elements(By.TAG_NAME, "a")  # Находим в блоке div все ссылки
        # p = [i.get_attribute('title') for i in self.href_list]
        #
        # for i in p:
        #     print(i)

        self.href_list = [i.get_attribute("href") for i in tqdm(self.href_list)]  # if
        # if self.href_list[0] == None:
        #     print("cсылки не полученны")
        # i.get_attribute("id") == "video-title"])  # забираем у всех ссылок атрибут href
        print(F"Полученно {len(self.href_list)} ссылок на видео. ")
        # print(self.href_list)
        return True

    def run(self, pipe_client=None):
        # '''Данная фукнция '''

        try:
            self.__driver = self.__create_driver()
            print("-" * 40)
            print(f"Идёт поиск всех видео канала {self.__chanal_url}.")
            if pipe_client:
                pipe_client.send(f"Идёт поиск всех видео канала {self.__chanal_url}.")

            self.__driver.get(self.__chanal_url)  # Получение страницы по указанному url
            time.sleep(10)
            self.__driver.find_element(By.XPATH,
                                       '//*[@id="tabsContent"]/tp-yt-paper-tab[2]/div').click()  # Переход на вкладку видео

            self.page_scroller()  # пролистываем страницу до конца
            self.page_source = self.__driver.page_source
            # with open("../data/htlm.txt", "w", encoding="utf-8") as file:
            #     file.write(self.page_source)

            check = self.get_hrefs()

            # print("set",set(self.href_list))
            # print(len(set(self.href_list)))

            if check == None:
                print("Повтор получения ссылок.")
                self.href_list.clear()
                check = self.get_hrefs()
            if pipe_client:
                pipe_client.send(f"Ссылки полученны")

            if self.__multiproc == False:
                self.pars_title_description()
            elif self.__multiproc == True:
                self.pars_title_description_multiprocess()
            elif self.__multiproc == None:
                pass

            self.lost_video_checker()
            if pipe_client:
                pipe_client.send(f"Запись результатов в таблицу Excel")
            if self.all_video_data_list and self.__multiproc != None:
                self.write_to_excel()
            # print(f"время: {datetime.datetime.now() - start}")

        except NoSuchElementException:
            print("Ошибка Youtube. Повторное подключение.")
            time.sleep(10)
            self.run_check += 1
            if self.__driver:
                self.__exit_driver()
            if self.run_check < 5:
                self.run()
            else:
                print("Ошибка подключения")
        except urllib3.exceptions.NewConnectionError:
            print("Странная ошибка.")
        except:
            raise
        finally:
            print("-" * 40)
            pipe_client.send(f"Конец")
            if self.__driver:
                self.__exit_driver()

    def lost_video_checker(self):
        lost = len(self.lost_href_list)
        print(f"Потери {lost}.")
        time.sleep(0.3)
        if lost:
            self.href_list = self.lost_href_list.copy()
            self.lost_href_list.clear()
            self.pars_title_description(extend_data=True)
            lost2 = len(self.lost_href_list)
            print(f"Сново потеряно:{lost2}.\nУдалось вернуть: {abs(lost2 - lost)}")
            lost = len(self.lost_href_list)
            if lost:
                self.href_list = self.lost_href_list.copy()
                self.lost_href_list.clear()
                self.pars_title_description(extend_data=True)
                lost2 = len(self.lost_href_list)
                print(f"Сново потеряно:{lost2}.\nУдалось вернуть: {abs(lost2 - lost)}")
                lost = len(self.lost_href_list)
                if lost:
                    self.href_list = self.lost_href_list.copy()
                    self.lost_href_list.clear()
                    self.pars_title_description(extend_data=True)
                    lost2 = len(self.lost_href_list)
                    print(f"Сново потеряно:{lost2}.\nУдалось вернуть: {abs(lost2 - lost)}")

    def pars_title_description(self, extend_data=False):
        print("Получение данных.")

        session = self.__create_session()
        self.lost_href_list.clear()
        if self.href_list:
            for href in tqdm(self.href_list):
                try:
                    response = session.get(href, timeout=10)
                    if response.ok:
                        start_json = response.text.find("var ytInitialData")
                        raw_json = response.text[
                                   start_json + len("var ytInitialData") + 2:response.text.find("</script>",
                                                                                                start_json) - 1].strip()  # обрезали лишнее, оставили только json

                        data = json_validator.pars_video_data(raw_json)
                        # self.writer_to_csv((data["name"], href, data["description"]))

                        match (data["name"], data["description"]):
                            case (str(), str()):
                                if extend_data == False:
                                    self.all_video_data_list.append((data["name"], href, data["description"]))
                                else:
                                    self.all_video_data_list.extend([(data["name"], href, data["description"])])
                            case (None, str()):
                                self.lost_href_list.append(href)
                            case (str(), None):
                                data["description"] = self.find_problem_discription(session, href)
                                if extend_data == False:
                                    self.all_video_data_list.append((data["name"], href, data["description"]))
                                else:
                                    self.all_video_data_list.extend([(data["name"], href, data["description"])])
                            case None, None:
                                self.lost_href_list.append(href)
                        # print("data[name]=", data['name'], "     data[description]=", data["description"])

                        # if data["name"] and data["description"]:
                        #     # print("data[name]=", data['name'], "     data[description]=", data["description"])
                        #     if extend_data == False:
                        #         self.all_video_data_list.append((data["name"], href, data["description"]))
                        #     else:
                        #         self.all_video_data_list.extend([(data["name"], href, data["description"])])
                        # else:
                        #     # print("data[name]=", data['name'], "     data[description]=", data["description"])
                        #     self.lost_href_list.append(href)

                    else:
                        print("Запрос на ютуб не удался")
                        self.lost_href_list.append(href)
                except urllib3.exceptions.ConnectTimeoutError:
                    print("Time out")
                    self.lost_href_list.append(href)
            # print(self.all_video_data)
            # self.write_to_excel()
        else:
            print("Список с ссылками пуст")
        # if session:
        #     session.close()

    @staticmethod
    def find_problem_discription(session, problem_url):
        try:
            response = session.get(problem_url)
            html = response.text
            data = None
            start = html.find("shortDescription")
            finish = html.find("isCrawlable")
            if start and finish and (finish > start):
                data = html[start + len("shortDescription") + 2:finish - 2]
                if data:
                    data = data.replace(r"\n\n", r"\n").replace(r"\n", "\n")
                    # print(data)
            return data
        except:
            return None

    @staticmethod
    def process(session, href, x: list, y: list, func):
        try:
            response = session.get(href, timeout=10)
            if response.ok:
                start = response.text.find("var ytInitialData")
                page_json = response.text[
                            start + len("var ytInitialData") + 2:response.text.find("</script>",
                                                                                    start) - 1].strip()  # обрезали лишнее,
                # оставили только json
                data = json_validator.pars_video_data(page_json)  # pars_video_data учитывает ошибки валидации

                match (data["name"], data["description"]):
                    case (str(), str()):
                        x.append((data["name"], href, data["description"]))
                    case (None, str()):
                        y.append(href)
                    case (str(), None):
                        data["description"] = func(session, href)
                        x.append((data["name"], href, data["description"]))
                    case None, None:
                        y.append(href)
                # print("data[name]=", data['name'], "     data[description]=", data["description"])

                # if data["name"] or data["description"]:
                #     x.append((data["name"], href, data["description"]))
                # else:
                #     y.append(href)
            else:
                print("Запрос на ютуб не удался")
                y.append(href)
        except urllib3.exceptions.ConnectTimeoutError:
            print("Time out")
            y.append(href)

    def pars_title_description_multiprocess(self, extend_data=False):
        print("Получение данных.")
        session = self.__create_session()
        self.lost_href_list.clear()
        if self.href_list:
            with multiprocessing.Manager() as manager:
                x = manager.list()
                y = manager.list()
                href_list = [(session, i, x, y, self.find_problem_discription) for i in self.href_list]

                with multiprocessing.Pool(multiprocessing.cpu_count() * 3) as pool:
                    pool.starmap(self.process, href_list)
                    pool.close()
                    pool.join()

                if extend_data == False:
                    self.all_video_data_list = list(x)
                else:
                    self.all_video_data_list.extend(list(x))
                self.lost_href_list = list(y)
        else:
            print("Список с ссылками пуст")
        # if session:
        #     session.close()

    @staticmethod
    def write_list(data_list, path):
        data_list = [i + "\n" for i in data_list]
        with open(f"{path}/потеряные ссылки.txt", "w", encoding="utf-8") as file:
            file.writelines(data_list)

    def write_to_excel(self):
        dir_name = self.__chanal_url.removeprefix('https://www.youtube.com/c/')
        dir_name = dir_name.replace("/", "_")
        time = datetime.datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
        if not os.path.exists(f"../data/{dir_name}"):
            os.mkdir(f"../data/{dir_name}")
        os.mkdir(f"../data/{dir_name}/{time}")
        work_book = openpyxl.Workbook()
        sheat = work_book["Sheet"]

        sheat["A1"] = "Название"
        sheat["B1"] = "Ссылка"
        sheat["C1"] = "Описание"

        row = 2
        for i in self.all_video_data_list:
            # print(i)
            # print(type(i))
            sheat.cell(row=row, column=1).value = i[0]
            sheat.cell(row=row, column=2).value = i[1]
            sheat.cell(row=row, column=3).value = i[2]
            row += 1
        # list_shets = work_book.sheetnames
        # print(list_shets)
        work_book.save(f"../data/{dir_name}/{time}/excel.xlsx")
        if self.lost_href_list:
            self.write_list(self.lost_href_list, f"../data/{dir_name}/{time}")

    # def create_new_csv(self):
    #     dir_name = self.__chanal_url.removeprefix('https://www.youtube.com/c/')
    #     if not os.path.exists(f"../data/{dir_name}"):
    #         os.mkdir(f"../data/{dir_name}")
    #     self.name_csv = f"../data/{dir_name}/{datetime.datetime.now().strftime('%d-%m-%Y')}.csv"
    #     with open(self.name_csv, "w", encoding="utf-8") as file:
    #         data = ("Название:", "Ссылка:", "Описание:")
    #         writer = csv.writer(file)
    #         writer.writerow(data)
    #
    # def writer_to_csv(self, data: tuple | list):
    #     with open(self.name_csv, "a", encoding="utf-8") as file:
    #         writer = csv.writer(file)
    #         writer.writerow(data)

    def __create_driver(self, driver_path: str = "../chromedriver.exe"):
        if os.path.exists(driver_path):
            driver_options = webdriver.ChromeOptions()
            driver_options.add_argument("--headless")

            driver_options.add_argument("--disable-blink-features=AutoControled")
            driver_options.add_argument(f"user-agent={self.__user_agent.chrome}")
            driver = webdriver.Chrome(executable_path=driver_path, chrome_options=driver_options)
            driver.implicitly_wait(10)
            return driver
        else:
            print("Веб драйвера не обнаружено, проверте правильно ли указан путь до него.")

    def __create_session(self):
        session = req.Session()
        session.headers.update({"user-name": self.__user_agent.chrome, "accept": "/"})
        return session

    def __exit_driver(self):
        if self.__driver:
            # self.__driver.close()
            self.__driver.quit()

    # def __del__(self):
    #     if self.__driver:
    #         self.__exit_driver()


def main():
    all_start = datetime.datetime.now()
    check_list = (
        "https://www.youtube.com/c/MeDallisTRoyale",
        # "https://www.youtube.com/с/Wylsacom",
        # "https://www.youtube.com/с/AcademeG",
        # "https://www.youtube.com/c/SuperCrastan",
        # "https://www.youtube.com/с/UCBUPvbjvN6Raly2FL14xoYw",
        # "https://www.youtube.com/c/JoeSpeen",
        # "https://www.youtube.com/c/UCfFkXO1Tfk6Je9O5h-dvLoA",
        #               "https://www.youtube.com/c/ZProgerIT",
        #               "https://www.youtube.com/c/Redlyy",
        #               "https://www.youtube.com/c/QuantumGames",
        #               "https://www.youtube.com/c/PhysicsisSimple"
        #               "https://www.youtube.com/c/gosha_dudar",
        #               "https://www.youtube.com/c/kuplinovplay"
    )
    with open("test_otchet.txt", "w") as file:
        for url in check_list:
            start = datetime.datetime.now()
            try:
                parser = YouTube_Parser(url, multiproc=True)
                parser.run()

                print("статистика:")
                print(len(parser.all_video_data_list))
                x = [i[0] for i in parser.all_video_data_list]
                y = [i[1] for i in parser.all_video_data_list]
                z = [i[2] for i in parser.all_video_data_list]
                print("Длины:", len(x), len(y), len(z))
                x = set(x)
                y = set(y)
                z = set(z)
                print("Длины:", len(x), len(y), len(z))

                file.write(f"{url} - true\n")
            except Exception as e:
                file.write(f"{url} - false\n")
                print(f"Ошибка {e}")
                raise
            except:
                file.write(f"{url} - false\n")
                print("Неожиданная ошибка")
                raise
            print(f"Время на парсинг {url} = {datetime.datetime.now() - start}")

    token = "5226592225:AAGuyEtD_FOotorITU45tTNOLWhEcR2htVA"
    chat_id = "488216212"
    res = req.get("https://api.telegram.org/bot" + token + "/sendMessage",
                  params={"chat_id": chat_id, "text": "программа завершена"})
    print("-" * 10)
    print(f"{res.ok=}")
    print("программа завершена!")
    print("затрачено времени:", datetime.datetime.now() - all_start)
    # for i in range(15):
    # x = YouTube_Parser("https://www.youtube.com/c/MeDallisTRoyale")
    # r = x.pars_title_description(("https://www.youtube.com/watch?v=yOvvtKRBzQU",))
    # del x


if __name__ == '__main__':
    # import pstats
    #
    # # cProfile.run("main()", sort="cumtime")
    # profiler = cProfile.Profile()
    # profiler.enable()
    # main()
    # profiler.disable()
    # stat = pstats.Stats(profiler).sort_stats("cumtime")
    # stat.strip_dirs()
    # stat.print_stats()
    # stat.dump_stats("1.prof")
    main()
